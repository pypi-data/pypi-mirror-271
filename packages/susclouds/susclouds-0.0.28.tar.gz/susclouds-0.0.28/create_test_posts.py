import pandas as pd
import os
import sys
import string
import itertools
from openpyxl import Workbook, load_workbook

def write_tab(wb, name, posts, brand):
    ws = wb.create_sheet(name)
    ws["a1"] = "Owner?"
    ws["b1"] = "Brand Owner"
    ws["c1"] = "Brand"
    ws["d1"] = "Platform"
    ws["e1"] = "URL"
    ws["f1"] = "Category"
    ws["g1"] = "Confirmed"
    ws["h1"] = "Interesting"
    ws["i1"] = "Timestamp"
    ws["j1"] = "Notes"
    ws["k1"] = "Clean Text"
    ws["l1"] = "Full Text"
    
    row = 2
    for post in posts:
        ws[f"c{row}"] = brand
        ws[f"l{row}"] = post
        row += 1
        
def append_rows(source,target):
    rows=1

    for row in source.rows:
        row_data=[]
        columns=1
        c=source.cell(row=rows,column=columns)
        if c.value is not None:
            while c.value is not None:
                c=source.cell(row=rows,column=columns)
                row_data.append(c.value)
                columns+=1
            
            c=target.append(row_data)
            rows+=1
        else:
            break

def main():
    source_file = "data/words for test posts.xlsx"
    word_bucket = []
    phrase_bucket = []
    stop_word_bucket = []
    brand_bucket = []
    posts = []
    tab_names = ["Emissions", "Materials", "Energy", "Waste", "Water", "Biodiversity", "T20 NZ",
                "1 Corp", "2 Corp", "3 Corp", "4 Corp", "5 Corp", "6 Corp", "7 Corp",
                "8 Corp", "9 Corp", "10 Corp", "11 Corp", "12 Corp", "13 Corp", "14 Corp",
                "15 Corp", "16 Corp", "17 Corp", "18 Corp", "19 Corp", "20 Corp"]
    tab_nos = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26]
    brand_names = ["nothing", "nothing", "nothing", "nothing", "nothing", "nothing", "nothing",
                    "Nestlé", "PespiCo", "Procter & Gamble", "Unilever", "AB InBev", "Mars Inc.",
                    "The Coca-Cola Company", "Danone", "Mondelēz", "The Heineken Company",
                    "Kraft Heinz", "Suntory", "Reckitt", "General Mills", "Diageo",
                    "Colgate-Palmolive", "Kimberly-Clark", "Ferrero", "Kellogg's", "Henkel"]
    
    words = pd.read_excel(source_file, sheet_name="Top words")
    phrases = pd.read_excel(source_file, sheet_name="Top phrases")
    stop_words = pd.read_excel(source_file, sheet_name="Top stop words")
    brands = pd.read_excel(source_file, sheet_name="Top brands")

    """
    create buckets of words, phrases, stop words and brands with each of those appearing with the
    frequency identified in the counts column.
    """
    for word, count in zip(words["words"],words["counts"]):
        i = 0
        while i < count:
            word_bucket.append(word)
            i += 1

    for word, count in zip(phrases["words"], phrases["counts"]):
        i = 0
        while i < count:
            phrase_bucket.append(word)
            i += 1

    for word, count in zip(stop_words["words"], stop_words["counts"]):
        i = 0
        while i < count:
            stop_word_bucket.append(word)
            i += 1

    for word, count in zip(brands["words"], brands["counts"]):
        i = 0
        while i < count:
            brand_bucket.append(word)
            i += 1

    """
    create a list of posts each containing a word, a stop word, phrase(x2) and some
    special characters
    """
    for word, phrase, stop_word, brand in zip(word_bucket, phrase_bucket, stop_word_bucket, brand_bucket):
        posts.append(f"{brand} {word} {phrase} {stop_word} {phrase} https://susmon.com, . : £")

    print(posts)

    wb_out = Workbook()
    wb_out_name = "data/Word Cloud Test Dataset.xlsx"

    """
    Fill the tabs of the test data set with posts
    """
    for tab, tab_no, brand in zip(tab_names, tab_nos, brand_names):
        post_index = tab_no
        tab_posts = []
        while post_index < len(posts):
            tab_posts.append(posts[post_index])
#            post_index += len(tab_nos)
            post_index += 1
        write_tab(wb_out, tab, tab_posts, brand)

    del wb_out["Sheet"]
    
    followers_book = load_workbook(source_file)
    follower_sheet = followers_book['Followers']
    new_followers_sheet = wb_out.create_sheet('Followers')
    append_rows(follower_sheet, new_followers_sheet)
    
    wb_out.save(wb_out_name)
    wb_out.close()

if __name__ == "__main__":
    main()
