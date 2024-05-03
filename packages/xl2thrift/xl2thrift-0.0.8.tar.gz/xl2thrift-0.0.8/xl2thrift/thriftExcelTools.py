import sys, os
import importlib
import datetime
import glob

import thrift
from thrift.transport import TTransport
from thrift.Thrift import TType
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

from .exceptions import EnumException, ThriftGeneratedModuleException, FileNotFoundException

def Log(s):
	verbose = globals()['verbose']
	if verbose:
	 	print(s)

def stringToName(s):
	name = s.split("--", 1)[0].strip().replace(" ", "")
	return name[0].lower() + name[1:]

class SheetColumnEntry:
	title = None
	attributeName = None
	columnNum = 0
	def __init__(self, cnum, topcell):
		self.title = topcell
		self.attributeName = stringToName(topcell)
		self.columnNum = cnum
	# 	self.dump()
	# def dump(self):
	# 	if globals()['verbose']:
	# 		print("SheetColumnEntry for column %d has attributeName %s (%s)" % (self.columnNum, self.attributeName, self.title))

class SheetInfo:
	title = None
	tablename = None
	columnInfo = None
	numcolumns = 0
	def __init__(self, sheet):
		self.title = sheet.title
		self.tablename = stringToName(self.title)
		self.columnInfo = {}

		#Log("New SheetInfo (%s) -> table %s" % (self.title, self.tablename))
		self.numcolumns = countActiveColumns(sheet)
		#print('--> new SheetInfo title:%s tablename:%s with %s columns' % (sheet.title, self.tablename, self.numcolumns))
		if self.numcolumns > 0:
			c = 1
			for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=self.numcolumns):
				for cell in row:
					topcell = cell.value
					entry = SheetColumnEntry(c, topcell)
					self.columnInfo[c] = entry
					c += 1
			checkdupes = {}
			for c in self.columnInfo:
				columnInfo = self.columnInfo[c]
				if columnInfo.attributeName in checkdupes:
					print('WARNING: Sheet <%s> has duplicate columns for <%s>' % (sheet.title, columnInfo.attributeName))
				checkdupes[columnInfo.attributeName] = 1
		self.dump()

	def dump(self):
		if (globals()['verbose']):
			print("++++++++++++++++++++++++++++++")
			print(("SheetInfo (%s) -> table %s" % (self.title, self.tablename)))
			for c in range(1,len(self.columnInfo) + 1):
				columnInfo = self.columnInfo[c]
				if columnInfo != None:
					print(("	column %d: cattribute name %s" % (c, columnInfo.attributeName)))
			print("++++++++++++++++++++++++++++++")

# type info for a field in the Data class
# if it's a container type then it's "of" value type.
class TypeEntry:
	type = None
	name = None
	mapIndexType = None
	valueType = None
	thriftSpec = None
	containerType = None
	def __init__(self, spec):
		if spec != None:
			self.thriftSpec = spec
			self.fieldNum = spec[0]
			self.containerType = spec[1]
			self.name = spec[2]
			if self.containerType == TType.MAP:
				self.mapIndexType = spec[3][0]
				if spec[3][3] == 'UTF8':
					# map of strings
					self.valueType = spec[3][3]
				elif spec[3][2] == TType.BOOL:
					# map of bool
					self.valueType = TType.BOOL
				else:
					self.valueType = spec[3][3][0]
			elif self.containerType == TType.LIST:
				if spec[3][1] == 'UTF8':
					##print(self.name + " is a list of strings")
					self.valueType = spec[3][1]
					#self.dump()
				elif spec[3][0] == TType.I32:
					self.valueType = spec[3][0]
				else:
					self.valueType = spec[3][1][0]
					# this constructs the object
					#new_object = self.valueType()
			elif self.containerType == TType.STRUCT:
				self.valueType = spec[3][0]
			self.dump()
	def dump(self):
		if (globals()['verbose']):
			print(("TypeEntry: %s in Data[%d], self.valueType %s" % (self.name, self.fieldNum, self.valueType)))

def makeTypeEntries(Data):
	typeEntries = {}
	for spec in Data.thrift_spec:
		typeEntry = TypeEntry(spec)
		if typeEntry != None:
			typeEntries[typeEntry.name] = typeEntry
	return typeEntries

def	countActiveColumns(sheet):
	count = 0
	# note max_row = 1 here
	#row = sheet[1]
	for row in sheet.iter_rows(min_row=1, min_col=1, max_row=1, max_col=1000):
		for cell in row:
			##print('countActiveColumns - content %s' %(cell.value))
			if cell.value == None:
				##print('countActiveColumns - count %s' %(count))
				return count
			count += 1

def getRowObject(sheetInfo, row, columnFieldSpec):
	enums = globals()['enums']
	if len(row) == 0 or row[0].value == None:
		return
	tablename = sheetInfo.tablename
	typeEntry = typeEntries[tablename]
	numcolumns = sheetInfo.numcolumns
	if typeEntry.valueType == 'UTF8':
		return row[0].value
	elif typeEntry.valueType == TType.BOOL:
		return row[0].value
	else:
		new_object = typeEntry.valueType()
		# populate list value object
		column = 1
		for cell in row:
			info = sheetInfo.columnInfo[column]
			value = cell.value
			if value != None:
				enumClass = None
				enumLookup = typeEntry.valueType.__name__ + '.' + info.attributeName
				if enumLookup in enums:
					enumClass = getEnumClass(enumLookup)
					# enumClassName = enums[enumLookup]
					# enumClass = getattr(ConfigModule, enumClassName)

				if column in columnFieldSpec:
					containerType = columnFieldSpec[column][1]
					if containerType == TType.STRUCT:
						fieldType = columnFieldSpec[column][3][0]
						if fieldType.__name__ == "DateTimeUTC":
							msec = int((value - datetime.datetime(1970,1,1)).total_seconds()) * 1000
							value = fieldType(msec)
					elif containerType == TType.LIST:
						# special case - allow ints to be lists of 1 int
						valueType = columnFieldSpec[column][3][0]
						if (valueType == TType.I32 or valueType == TType.DOUBLE) and not hasattr(value, 'split'):
							value = [ value ]
						else:
							value = value.split(';')[0]
							if enumClass != None:
								valueStrings = value.split(',')
								value = []
								for v in valueStrings:
									value.append(enumClass._NAMES_TO_VALUES[v])
							elif valueType == TType.I32:
								valueStrings = value.split(',')
								value = []
								for v in valueStrings:
									value.append(int(v))
							elif valueType == TType.DOUBLE:
								valueStrings = value.split(',')
								value = []
								for v in valueStrings:
									value.append(float(v))
							elif valueType == TType.BOOL:
								valueStrings = value.split(',')
								value = []
								for v in valueStrings:
									if v.lower()[0] == 't' or v == '1':
										b = True
									elif v.lower()[0] == 'f' or v == '0':
										b = False
									value.append(b)
							else:
								value = value.split(',')
								for v in value:
									if v != v.strip():
										print(("WARNING: the string '%s' has leading/trailing spaces" % (v)))

				elif enumClass != None:
					try:
						value = enumClass._NAMES_TO_VALUES[value]
					except:
						raise EnumException('ERROR: undefined enum value <%s>' % (value))
			# get member for column
			info = sheetInfo.columnInfo[column]
			setattr(new_object, info.attributeName, value)
			column += 1
			if column > numcolumns:
				break
		new_object.validate()
	return new_object

def parseSheet(sheet):
	sheetInfo = SheetInfo(sheet)
	numcolumns = sheetInfo.numcolumns
	enums = globals()['enums']
	#sheetInfo.dump()
	tablename = stringToName(sheet.title)
	if not tablename in typeEntries:
		Log('Ignore table <%s>' % (tablename))
	if tablename in typeEntries:
		Log("parseSheet " + sheet.title + " with tablename " + tablename)
		typeEntry = typeEntries[tablename]
		##print('typeEntry name %s typename %s' % (typeEntry.name, typeEntry.containerType.__name__))
		table = getattr(Data, tablename)
		if table == None:
			##Log("created table " + tablename)
			if typeEntry.containerType == TType.MAP:
				setattr(Data, tablename, {})
				table = getattr(Data, tablename)
			elif typeEntry.containerType == TType.LIST:
				setattr(Data, tablename, [])
				table = getattr(Data, tablename)
			elif typeEntry.containerType == TType.STRUCT:
				setattr(Data, tablename, typeEntry.valueType())
		# Kludge - we only need this for structs but we create it for tables too
		new_object = getattr(Data, tablename)

		# map columns to fieldNums. getting the field type involves traversing a complicated thrift structure
		columnFieldSpec = {}
		dataLineThriftSpec = typeEntry.thriftSpec
		Log("Field %s" % (dataLineThriftSpec[2]))
		rowFieldType = None
		if typeEntry.containerType == TType.MAP:
			if dataLineThriftSpec[3][3] == 'UTF8':
				rowFieldType = dataLineThriftSpec[3][3]
				Log('map of string to string rowFieldType %s' % (rowFieldType))
			elif hasattr(dataLineThriftSpec[3][3], "__iter__"):
				rowFieldType = dataLineThriftSpec[3][3][0]
				Log('map of string to struct rowFieldType %s' % (rowFieldType))
		elif typeEntry.containerType == TType.LIST:
			if dataLineThriftSpec[3][1] == 'UTF8':
				rowFieldType = dataLineThriftSpec[3][1]
			elif hasattr(dataLineThriftSpec[3][1], "__iter__"):
				rowFieldType = dataLineThriftSpec[3][1][0]
		elif typeEntry.containerType == TType.STRUCT and hasattr(dataLineThriftSpec[3], "__iter__"):
			rowFieldType = dataLineThriftSpec[3][0]

		if rowFieldType == 'UTF8':
			Log("string map rowFieldType %s" % (rowFieldType))
			columnFieldSpec[1] = 'UTF8'
		elif rowFieldType != None:
			Log("rowFieldType %s" % (rowFieldType))
			for c in range(1,len(sheetInfo.columnInfo) + 1):
				columnInfo = sheetInfo.columnInfo[c]
				if columnInfo != None:
					#Log("parseSheet: %d: column %d attribute name %s" % (c, columnInfo.columnNum, columnInfo.attributeName))
					for fieldSpec in rowFieldType.thrift_spec:
						if fieldSpec != None:
							if fieldSpec[2] == columnInfo.attributeName and hasattr(fieldSpec[3], "__iter__"):
								#print('Found it: columnInfo.attributeName %s' % (columnInfo.attributeName))
								columnFieldSpec[c] = fieldSpec

		rows = sheet.rows
		rownum = 1
		for row in rows:
			if rownum == 1:
				rownum = rownum
			elif row == () or row[0].value == None:
				break
			elif typeEntry.containerType == TType.STRUCT:
				attrname = stringToName(row[0].value)
				attrvalue = row[1].value
				rowFieldType = dataLineThriftSpec[3][0]
# 				print('rowFieldType %s, typeEntry.containerType %s' % (rowFieldType, typeEntry.containerType))
				for fieldSpec in rowFieldType.thrift_spec:
					if fieldSpec != None:
						if fieldSpec[2] == attrname:
							valueType = fieldSpec[1]
							# structs can contain lists as values, but only of i32, double and string
							if valueType == TType.LIST:
								subValueType = fieldSpec[3][0]
								valueStrings = attrvalue.split(',')
								value = []
								for v in valueStrings:
									if subValueType == TType.I32:
										enumLookup = typeEntry.valueType.__name__ + '.' + fieldSpec[2]
										if enumLookup in enums:
											enumClass = getEnumClass(enumLookup)
											# enumClassName = enums[enumLookup]
											# enumClass = getattr(ConfigModule, enumClassName)
											value.append(enumClass._NAMES_TO_VALUES[v])
										else:
											value.append(int(v))
									elif subValueType == TType.DOUBLE:
										value.append(float(v))
									else:
										value.append(v)
								setattr(new_object, attrname, value)
							else:
								setattr(new_object, attrname, attrvalue)
			elif typeEntry.containerType == TType.MAP:
				if typeEntry.valueType == 'UTF8':
					objectIndex = row[0].value
					if objectIndex != objectIndex.strip():
						print(("WARNING: index <%s> in table <%s> contains leading and/or trailing spaces" % (objectIndex, tablename)))
					attrvalue = row[1].value
					if objectIndex in table:
						print(("WARNING: Duplicate entry in table <%s> at index <%s>" % (tablename, objectIndex)))
					table[objectIndex] = attrvalue
				elif typeEntry.valueType == TType.BOOL:
					objectIndex = row[0].value
					attrvalue = row[1].value
					table[objectIndex] = attrvalue
				else:
					new_object = getRowObject(sheetInfo, row, columnFieldSpec)
					indexAttributeName = sheetInfo.columnInfo[1].attributeName
					objectIndex = getattr(new_object, indexAttributeName)
					if objectIndex in table:
						print(("WARNING: Duplicate entry in table <%s> at index <%s>" % (tablename, objectIndex)))
					table[objectIndex] = new_object
					new_object.validate()
			elif typeEntry.containerType == TType.LIST:
				new_object = getRowObject(sheetInfo, row, columnFieldSpec)
				table.append(new_object)
			rownum += 1

		if typeEntry.containerType == TType.STRUCT:
			table = new_object
			new_object.validate()

def parseEnums(path):
	try:
		with open(path) as myfile:
			for line in myfile:
				name, var = line.split("=")
				enums[name.strip()] = var.strip()
	except:
		raise FileNotFoundException('Enums file <%s> not found' % (path))
	return enums

def getEnumClass(enumLookup):
	ConfigModule = globals()['ConfigModule']
	enumClassName = enums[enumLookup]
	enumClass = getattr(ConfigModule, enumClassName)
	return enumClass

def parseWorkbook(workbook):
	for sheet in workbook:
		parseSheet(sheet)

def parseFile(path):
	Log('========== Workbook %s ============' % (path))
	workbook = load_workbook(path, read_only = True, data_only = True)
	parseWorkbook(workbook)

def usage():
	print('--gen_py should point to your thrift-generated python source folder')
	print('    i.e, the folder generated by `thrift --gen py <thriftfile>')
	print('--namespace is the namespace in your thrift file')
	print('--class_name is the class in your thrift file that contains all the data')
	print('--verbose for more output')
	print('--enums coming soon')

# Globals
typeEntries = {}
Data = {}
enums = {}
args = {}

def list_matching_files(glob_patterns):
    matched_files = []
    for pattern in glob_patterns.split():
        matched_files.extend(glob.glob(pattern))
    return matched_files

def convertXlsxToThrift(patterns, namespace, thrift_protocol, gen_py, class_name='Data', output='config.bin', enums_path=None, release=False, verbose=False):
	globals()['verbose'] = verbose
	sys.path.append(gen_py)
	ConfigModule = {}
	try:
		ConfigModule = importlib.import_module('%s.ttypes' % (namespace))
	except:
		print('Failed to load thrift-generated module (%s.ttypes)' % (namespace))
		print('Check that the namespace in your thrift file matches the --namespace arg (%s)' % (namespace))
		print('and that the --gen_py arg (%s) points to the correct folder' % (gen_py))
		raise ThriftGeneratedModuleException('Failed to load thrift-generated module in <%s> called (%s.ttypes)' % (gen_py, namespace))

	sys.path.remove(gen_py)
	Log(ConfigModule)
	globals()['ConfigModule'] = ConfigModule

	try:
		dataClass = getattr(ConfigModule, class_name)
	except:
		print('Please make sure that your --class_name arg refers to the correct class in your thrift file (and in the source code in --gen_py)')
		raise FileNotFoundException('Failed create an instance of class (%s) specified in the --class_name arg' % (class_name))

	globals()['Data'] = dataClass()
	Data = globals()['Data']

	sheetInfo = None

	if enums_path:
		globals()['enums'] = parseEnums(enums_path)
	globals()['typeEntries'] = makeTypeEntries(Data)

	filenames = list_matching_files(patterns)
	for path in filenames:
		file_ = os.path.basename(path)
		if file_[:2] != "~$" and file_[:2] != "--" and file_.lower().endswith(".xlsx"):
			if not file_.lower().endswith("!.xlsx"):
				parseFile(path)

	# Debug sheets are added last so that they can overwrite 
	for path in filenames:
		file_ = os.path.basename(path)
		if file_[:2] != "~$" and file_[:2] != "--" and file_.lower().endswith(".xlsx"):
			if not release and file_.endswith('!.xlsx'):
				parseFile(path)

	# for pattern in patterns:
	# 	for path in glob.glob(pattern, recursive=True):
	# 		file_ = os.path.basename(path)
	# 		if file_[:2] != "~$" and file_[:2] != "--" and file_.lower().endswith(".xlsx"):
	# 			if not file_.lower().endswith("!.xlsx"):
	# 				parseFile(path)

	# Debug sheets are added last so that they can overwrite 
	# for pattern in patterns:
	# 	for path in glob.glob(pattern, recursive=True):
	# 		file_ = os.path.basename(path)
	# 		if file_[:2] != "~$" and file_[:2] != "--" and file_.lower().endswith(".xlsx"):
	# 			if not release and file_.endswith('!.xlsx'):
	# 				parseFile(path)

	transport = TTransport.TMemoryBuffer()
	ThriftProtocol = getattr(importlib.import_module("thrift.protocol.%s" % (thrift_protocol)), thrift_protocol)
	protocol = ThriftProtocol(transport)
	Data.schemaVersionId = 1
	Data.write(protocol)

	# Write output file
	configPath = output
	buf = transport.getvalue()
	with open(configPath, 'wb') as f:
		f.write(buf)

	# Read and validate
	buf = None

	with open(configPath, 'rb') as f:
		buf = f.read()

	transport = TTransport.TMemoryBuffer(buf)
	protocol = ThriftProtocol(transport)
	Data = None
	Data = dataClass()
	Data.read(protocol)
	return Data.validate()
